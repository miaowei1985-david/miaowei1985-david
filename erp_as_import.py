#!/usr/bin/env python3
"""
京东售后数据导入数据库，并关联到财务匹配表
优化版：批量查询替代逐条查询
"""
import openpyxl
import os
import sqlite3
import logging
logging.basicConfig(filename="/tmp/erp_as_import.log", level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "uploads", "after_sales_upload.xlsx")
DB_PATH = os.path.join(BASE_DIR, "erp_all.db")


def main():
    print("📊 读取售后Excel...", flush=True)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    # 第2行是表头
    cols = []
    for cell in ws[2]:
        if cell.value:
            cols.append(cell.value)
        else:
            cols.append(f"col_{cell.column}")

    # 从第3行开始读数据
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:  # 服务单号存在
            record = {}
            for i, c in enumerate(cols):
                if i < len(row):
                    record[c] = row[i]
                else:
                    record[c] = None
            rows.append(record)

    wb.close()
    print(f"  售后记录: {len(rows)} 条", flush=True)

    # 收集所有不重复的tid
    all_tids = set()
    for r in rows:
        tid = str(r.get("订单号", "") or "").strip()
        if tid:
            all_tids.add(tid)
    all_tids_list = sorted(all_tids)
    print(f"  唯一订单号: {len(all_tids_list)} 个", flush=True)

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")

    # 批量查 finance_matched
    print("  批量查询财务数据...", flush=True)
    cache_finance = {}
    placeholders = ",".join("?" for _ in all_tids_list)
    for frow in conn.execute(
        f"SELECT tids, shop_name, warehouse_name, 总收入, 总支出, 净额, 结算状态 FROM finance_matched WHERE tids IN ({placeholders})",
        all_tids_list
    ):
        cache_finance[frow[0]] = frow[1:]

    # 批量查三张ERP表
    print("  批量查询ERP订单...", flush=True)
    cache_erp = {}
    for tbl in ["erp_wait_check", "erp_wait_send_self", "erp_finished"]:
        exists = conn.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{tbl}'").fetchone()
        if not exists:
            continue
        for erow in conn.execute(
            f"SELECT tid, shopName, warehouseName, paid FROM {tbl} WHERE tid IN ({placeholders})",
            all_tids_list
        ):
            if erow[0] not in cache_erp:
                cache_erp[erow[0]] = (erow[1], erow[2], erow[3])

    conn.close()
    print(f"  ERP匹配: {len(cache_erp)}, 财务匹配: {len(cache_finance)}", flush=True)

    # 写入数据库
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS after_sales")

    c.execute('''CREATE TABLE after_sales (
        service_no TEXT, tid TEXT, customer_expect TEXT, service_status TEXT,
        primary_reason TEXT, secondary_reason TEXT, description TEXT,
        return_method TEXT, apply_time TEXT, order_type TEXT,
        outbound_status TEXT, order_status TEXT, account TEXT,
        product_code TEXT, product_name TEXT, product_amount REAL,
        product_qty INT, refund_amount REAL,
        audit_result TEXT, audit_time TEXT, audit_opinion TEXT,
        auditor TEXT, logistics_no TEXT, shipping_cost REAL,
        express_company TEXT, result TEXT, handler TEXT,
        auto_audit_type TEXT, first_audit_time TEXT, first_process_time TEXT,
        duration_hours REAL,
        shop_name TEXT, warehouse_name TEXT, erp_paid REAL,
        finance_income REAL, finance_expense REAL, finance_net REAL,
        finance_status TEXT
    )''')

    print("  插入数据...", flush=True)
    placeholders38 = ",".join(["?"] * 38)
    batch = []
    for r in rows:
        tid = str(r.get("订单号", "") or "").strip()
        erp = cache_erp.get(tid)
        fin = cache_finance.get(tid)

        batch.append((
            r.get("服务单号", ""),
            tid,
            r.get("客户期望", ""),
            r.get("服务单状态", ""),
            r.get("一级申请原因", ""),
            r.get("二级申请原因", ""),
            r.get("客户问题描述", ""),
            r.get("返回方式", ""),
            r.get("售后申请时间", ""),
            r.get("订单类型", ""),
            r.get("出库状态", ""),
            r.get("申请时订单状态", ""),
            r.get("下单账号", ""),
            r.get("商品编号", ""),
            r.get("商品名称", ""),
            float(r.get("商品金额", 0) or 0),
            int(r.get("商品数量", 0) or 0),
            float(r.get("退款金额", 0) or 0),
            r.get("审核结果", ""),
            r.get("审核时间", ""),
            r.get("审核意见", ""),
            r.get("审核人姓名", ""),
            r.get("运单号", ""),
            float(r.get("运费金额", 0) or 0),
            r.get("快递公司", ""),
            r.get("处理结果", ""),
            r.get("处理人", ""),
            r.get("自动审核类型", ""),
            r.get("商家首次审核时间", ""),
            r.get("商家首次处理时间", ""),
            float(r.get("售后整体时长(H)", 0) or 0),
            (erp[0] if erp else "") or (fin[0] if fin else "") or ("榴愿时刻工厂店" if r.get("订单类型") == "POPSOP" else ""),
            erp[1] if erp else "",
            float(erp[2] if erp and erp[2] else 0),
            fin[1] if fin else 0,
            fin[2] if fin else 0,
            fin[3] if fin else 0,
            fin[4] if fin else "",
        ))
        if len(batch) >= 1000:
            c.executemany(f'INSERT INTO after_sales VALUES ({placeholders38})', batch)
            batch = []
    if batch:
        c.executemany(f'INSERT INTO after_sales VALUES ({placeholders38})', batch)

    conn.commit()
    c.execute("CREATE INDEX IF NOT EXISTS idx_as_tid ON after_sales(tid)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_as_status ON after_sales(service_status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_as_reason ON after_sales(primary_reason)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_as_outbound ON after_sales(outbound_status)")
    conn.commit()
    conn.close()

    print("✅ 售后数据已入库", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.exception("Script failed: %s", e)
        raise
